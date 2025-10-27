"""
Management command to import provider data from Excel files for specific regional centers.
Specifically designed for Pasadena and San Gabriel/Pomona regional center provider lists.

Usage:
    python manage.py import_regional_center_providers --file data/Pasadena_Provider_List.xlsx --regional-center "Pasadena"
    python manage.py import_regional_center_providers --file data/San_Gabriel_Pomona_Provider_List.xlsx --regional-center "San Gabriel/Pomona"
"""

import openpyxl
import os
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.conf import settings
from locations.models import ProviderV2, RegionalCenter
import requests
import time


class Command(BaseCommand):
    help = "Import providers from Excel files and associate with regional centers"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=str,
            required=True,
            help="Path to Excel file (e.g., data/Pasadena_Provider_List.xlsx)",
        )
        parser.add_argument(
            "--regional-center",
            type=str,
            required=False,
            help="Regional center name (e.g., 'Pasadena' or 'San Gabriel/Pomona')",
        )
        parser.add_argument(
            "--area",
            type=str,
            required=False,
            help="Area/city name for providers without specific regional center (e.g., 'Pasadena')",
        )
        parser.add_argument(
            "--clear-existing",
            action="store_true",
            help="Clear existing providers for this regional center before importing",
        )
        parser.add_argument(
            "--geocode",
            action="store_true",
            help="Geocode addresses using Mapbox API (requires MAPBOX_ACCESS_TOKEN env var)",
        )
        parser.add_argument(
            "--sheet",
            type=str,
            default=None,
            help="Specific sheet name to import (default: first sheet)",
        )

    def handle(self, *args, **options):
        file_path = options["file"]
        regional_center_name = options.get("regional_center")
        area_name = options.get("area")

        # Check if file exists
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"File not found: {file_path}"))
            return

        # Find or get the regional center (optional now)
        regional_center = None
        if regional_center_name:
            try:
                regional_center = RegionalCenter.objects.get(regional_center__icontains=regional_center_name)
                self.stdout.write(self.style.SUCCESS(f"Found regional center: {regional_center.regional_center}"))
            except RegionalCenter.DoesNotExist:
                self.stdout.write(self.style.ERROR(f"Regional center not found: {regional_center_name}"))
                self.stdout.write("Available regional centers:")
                for rc in RegionalCenter.objects.all().distinct('regional_center'):
                    self.stdout.write(f"  - {rc.regional_center}")
                return
            except RegionalCenter.MultipleObjectsReturned:
                regional_center = RegionalCenter.objects.filter(regional_center__icontains=regional_center_name).first()
                self.stdout.write(self.style.WARNING(f"Multiple matches found, using: {regional_center.regional_center}"))

        # Use area name if provided
        if area_name and not regional_center_name:
            self.stdout.write(self.style.SUCCESS(f"Importing providers for area: {area_name}"))
        elif not regional_center and not area_name:
            self.stdout.write(self.style.ERROR("Please provide either --regional-center or --area"))
            return

        if options["clear_existing"]:
            # Note: We don't have a direct FK to regional center, so we'd need to identify by area
            self.stdout.write(self.style.WARNING("Clear existing is not implemented yet for providers"))

        # Import the providers
        self.import_providers(file_path, regional_center, area_name or (regional_center.city if regional_center else ""), options)

    def import_providers(self, file_path, regional_center, area_name, options):
        """Import providers from Excel file"""
        self.stdout.write(f"\nImporting providers from {file_path}...")
        if regional_center:
            self.stdout.write(f"Regional Center: {regional_center.regional_center}")
        self.stdout.write(f"Area: {area_name}\n")

        # Open workbook
        try:
            workbook = openpyxl.load_workbook(file_path)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error opening Excel file: {e}"))
            return

        # Get the sheet
        if options["sheet"]:
            try:
                sheet = workbook[options["sheet"]]
            except KeyError:
                self.stdout.write(self.style.ERROR(f"Sheet not found: {options['sheet']}"))
                self.stdout.write(f"Available sheets: {', '.join(workbook.sheetnames)}")
                return
        else:
            sheet = workbook.active
            self.stdout.write(f"Using sheet: {sheet.title}\n")

        # Read the header row to understand the columns
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value.strip() if cell.value else "")

        self.stdout.write(f"Columns found: {', '.join(headers)}\n")

        # Map common column names (case-insensitive)
        column_map = self.map_columns(headers)
        self.stdout.write(f"Column mapping: {column_map}\n")

        created_count = 0
        updated_count = 0
        error_count = 0

        # Process each row (skip header)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Build a dict from the row
                row_data = {}
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        row_data[headers[idx]] = value

                # Extract provider data
                provider_name = self.get_value(row_data, column_map.get("name"))

                # Skip empty rows
                if not provider_name or provider_name.strip() == "":
                    continue

                # Parse the data
                provider_data = self.parse_provider_data(row_data, column_map, area_name)

                # Geocode if requested
                if options["geocode"] and provider_data.get("address"):
                    coords = self.geocode_address(provider_data["address"])
                    if coords:
                        provider_data["latitude"] = coords["latitude"]
                        provider_data["longitude"] = coords["longitude"]
                        time.sleep(0.2)  # Rate limiting for Mapbox API

                # Create or update provider
                provider, created = ProviderV2.objects.update_or_create(
                    name=provider_data["name"],
                    address=provider_data.get("address", ""),
                    defaults=provider_data
                )

                if created:
                    created_count += 1
                    self.stdout.write(f"  ✓ Created: {provider.name}")
                else:
                    updated_count += 1
                    self.stdout.write(f"  ↻ Updated: {provider.name}")

            except Exception as e:
                error_count += 1
                self.stdout.write(self.style.ERROR(f"  ✗ Error on row {row_num}: {e}"))

        # Summary
        self.stdout.write("\n" + "="*50)
        self.stdout.write(self.style.SUCCESS(f"Import complete!"))
        self.stdout.write(f"  Created: {created_count}")
        self.stdout.write(f"  Updated: {updated_count}")
        self.stdout.write(f"  Errors:  {error_count}")
        self.stdout.write("="*50 + "\n")

    def map_columns(self, headers):
        """Map Excel columns to provider fields"""
        column_map = {}

        for idx, header in enumerate(headers):
            header_lower = header.lower() if header else ""

            # Provider Name
            if any(x in header_lower for x in ["provider", "name", "organization"]):
                column_map["name"] = header

            # Address
            elif any(x in header_lower for x in ["address", "location", "street"]):
                column_map["address"] = header

            # Phone
            elif any(x in header_lower for x in ["phone", "telephone", "tel"]):
                column_map["phone"] = header

            # Services
            elif any(x in header_lower for x in ["service", "therapy", "treatment"]):
                column_map["services"] = header

            # Insurance
            elif any(x in header_lower for x in ["insurance", "payment", "funding"]):
                column_map["insurance"] = header

            # Notes
            elif any(x in header_lower for x in ["note", "comment", "description"]):
                column_map["notes"] = header

            # Email
            elif any(x in header_lower for x in ["email", "e-mail"]):
                column_map["email"] = header

            # Website
            elif any(x in header_lower for x in ["website", "web", "url"]):
                column_map["website"] = header

        return column_map

    def get_value(self, row_data, column_name):
        """Safely get value from row data"""
        if not column_name:
            return None
        value = row_data.get(column_name)
        if value is None or (isinstance(value, str) and value.strip() == ""):
            return None
        return str(value).strip() if value else None

    def parse_provider_data(self, row_data, column_map, area_name):
        """Parse provider data from row"""
        data = {
            "name": self.get_value(row_data, column_map.get("name", "")),
            "phone": self.get_value(row_data, column_map.get("phone")),
            "email": self.get_value(row_data, column_map.get("email")),
            "website": self.get_value(row_data, column_map.get("website")),
            "address": self.get_value(row_data, column_map.get("address", "")),
            "verified": False,
        }

        # Parse services into therapy_types
        services_text = self.get_value(row_data, column_map.get("services"))
        if services_text:
            therapy_types = self.parse_therapy_types(services_text)
            if therapy_types:
                data["therapy_types"] = therapy_types

        # Parse insurance
        insurance_text = self.get_value(row_data, column_map.get("insurance"))
        if insurance_text:
            insurance_list, accepts_flags = self.parse_insurance(insurance_text)
            if insurance_list:
                data["insurance_accepted"] = ", ".join(insurance_list)
            data.update(accepts_flags)

        # Parse notes into description
        notes = self.get_value(row_data, column_map.get("notes"))
        if notes:
            data["description"] = notes

        # Set area-specific data
        # Store in specific_areas_served
        if area_name:
            data["specific_areas_served"] = [area_name]

        # Default to LA County
        data["serves_la_county"] = True
        data["center_based_services"] = True

        return data

    def parse_therapy_types(self, services_text):
        """Parse services text into therapy types array"""
        therapy_types = []
        services_lower = services_text.lower()

        # Map service keywords to therapy types
        therapy_mapping = {
            "aba": "ABA therapy",
            "applied behavior": "ABA therapy",
            "speech": "Speech therapy",
            "occupational": "Occupational therapy",
            "ot": "Occupational therapy",
            "physical": "Physical therapy",
            "pt": "Physical therapy",
            "feeding": "Feeding therapy",
            "parent": "Parent child interaction therapy/parent training behavior management",
        }

        for keyword, therapy_type in therapy_mapping.items():
            if keyword in services_lower and therapy_type not in therapy_types:
                therapy_types.append(therapy_type)

        return therapy_types if therapy_types else None

    def parse_insurance(self, insurance_text):
        """Parse insurance text into list and acceptance flags"""
        insurance_list = []
        accepts_flags = {
            "accepts_insurance": False,
            "accepts_regional_center": False,
            "accepts_private_pay": False,
        }

        insurance_lower = insurance_text.lower()

        # Check for Regional Center
        if "regional center" in insurance_lower or "rc" in insurance_lower:
            insurance_list.append("Regional Center")
            accepts_flags["accepts_regional_center"] = True

        # Check for Private Pay
        if "private" in insurance_lower or "self pay" in insurance_lower or "cash" in insurance_lower:
            insurance_list.append("Private Pay")
            accepts_flags["accepts_private_pay"] = True

        # Check for common insurance types
        insurance_keywords = {
            "medi-cal": "Medi-Cal",
            "medicaid": "Medicaid",
            "medicare": "Medicare",
            "blue cross": "Blue Cross",
            "blue shield": "Blue Shield",
            "anthem": "Anthem",
            "aetna": "Aetna",
            "cigna": "Cigna",
            "kaiser": "Kaiser Permanente",
            "united": "United Healthcare",
            "health net": "Health Net",
            "molina": "Molina",
            "la care": "L.A. Care",
            "l.a. care": "L.A. Care",
        }

        for keyword, insurance_name in insurance_keywords.items():
            if keyword in insurance_lower and insurance_name not in insurance_list:
                insurance_list.append(insurance_name)
                accepts_flags["accepts_insurance"] = True

        return insurance_list, accepts_flags

    def geocode_address(self, address):
        """Geocode address using Mapbox Geocoding API"""
        mapbox_token = os.environ.get("MAPBOX_ACCESS_TOKEN")

        if not mapbox_token:
            self.stdout.write(self.style.WARNING("MAPBOX_ACCESS_TOKEN not set, skipping geocoding"))
            return None

        # Clean up address
        address = address.replace("\n", ", ").strip()

        # Mapbox Geocoding API
        url = f"https://api.mapbox.com/geocoding/v5/mapbox.places/{address}.json"
        params = {
            "access_token": mapbox_token,
            "country": "US",
            "limit": 1,
        }

        try:
            response = requests.get(url, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()

            if data.get("features"):
                feature = data["features"][0]
                longitude, latitude = feature["geometry"]["coordinates"]
                return {
                    "latitude": Decimal(str(latitude)),
                    "longitude": Decimal(str(longitude)),
                }
        except Exception as e:
            self.stdout.write(self.style.WARNING(f"Geocoding failed for {address}: {e}"))

        return None
