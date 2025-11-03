"""Excel file parsing utilities for provider data imports."""
import openpyxl
from typing import Dict, List, Optional, Any


class ExcelParser:
    """Handles parsing of Excel files for provider data."""

    def __init__(self, file_path: str, sheet_name: Optional[str] = None):
        """
        Initialize Excel parser.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Optional specific sheet name to use
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook = None
        self.sheet = None
        self.headers: List[str] = []

    def open(self) -> bool:
        """
        Open the Excel workbook and sheet.
        
        Returns:
            True if successful, False otherwise
        """
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            
            if self.sheet_name:
                if self.sheet_name not in self.workbook.sheetnames:
                    available = ', '.join(self.workbook.sheetnames)
                    raise ValueError(
                        f"Sheet '{self.sheet_name}' not found. "
                        f"Available sheets: {available}"
                    )
                self.sheet = self.workbook[self.sheet_name]
            else:
                self.sheet = self.workbook.active
            
            # Read headers
            self.headers = [
                cell.value.strip() if cell.value else ""
                for cell in self.sheet[1]
            ]
            
            return True
            
        except Exception as e:
            raise Exception(f"Error opening Excel file: {e}")

    def get_sheet_name(self) -> str:
        """Get the name of the active sheet."""
        return self.sheet.title if self.sheet else ""

    def get_headers(self) -> List[str]:
        """Get column headers from the sheet."""
        return self.headers

    def iter_rows(self, start_row: int = 2) -> List[Dict[str, Any]]:
        """
        Iterate through rows and yield data as dictionaries.
        
        Args:
            start_row: Row number to start from (default: 2, skipping header)
            
        Yields:
            Dictionary mapping header names to cell values
        """
        if not self.sheet:
            return

        for row_num, row in enumerate(
            self.sheet.iter_rows(min_row=start_row, values_only=True),
            start=start_row
        ):
            row_data = {}
            for idx, value in enumerate(row):
                if idx < len(self.headers):
                    row_data[self.headers[idx]] = value
            
            yield row_num, row_data

    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()


class ColumnMapper:
    """Maps Excel columns to provider fields based on header names."""

    COLUMN_MAPPINGS = {
        "name": ["provider", "name", "organization"],
        "address": ["address", "location", "street"],
        "phone": ["phone", "telephone", "tel"],
        "services": ["service", "therapy", "treatment"],
        "insurance": ["insurance", "payment", "funding"],
        "notes": ["note", "comment", "description"],
        "email": ["email", "e-mail"],
        "website": ["website", "web", "url"],
    }

    @classmethod
    def map_columns(cls, headers: List[str]) -> Dict[str, str]:
        """
        Map Excel column headers to standardized field names.
        
        Args:
            headers: List of column header strings from Excel
            
        Returns:
            Dictionary mapping field names to actual column names
        """
        column_map = {}

        for header in headers:
            if not header:
                continue
                
            header_lower = header.lower()

            for field_name, keywords in cls.COLUMN_MAPPINGS.items():
                if any(keyword in header_lower for keyword in keywords):
                    # Only map if not already mapped
                    if field_name not in column_map:
                        column_map[field_name] = header
                    break

        return column_map

    @staticmethod
    def get_value(row_data: Dict[str, Any], column_name: Optional[str]) -> Optional[str]:
        """
        Safely extract and clean value from row data.
        
        Args:
            row_data: Dictionary of row data
            column_name: Name of column to extract
            
        Returns:
            Cleaned string value or None
        """
        if not column_name:
            return None
            
        value = row_data.get(column_name)
        
        if value is None:
            return None
            
        if isinstance(value, str) and value.strip() == "":
            return None
            
        return str(value).strip() if value else None

